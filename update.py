"""
Samsung Pre-order Dashboard — Data Updater

Usage:
  1. Download the latest Excel from Shopl
  2. Drop it into this folder
  3. Run: python update.py

  It will find the newest .xlsx file, parse it, update data.json, and deploy to GitHub Pages.
"""
import openpyxl, json, glob, os, subprocess, sys
from datetime import datetime

# Find the newest xlsx file
xlsx_files = glob.glob(os.path.join(os.path.dirname(__file__) or '.', '*.xlsx'))
if not xlsx_files:
    print('No .xlsx files found in this folder.')
    sys.exit(1)

newest = max(xlsx_files, key=os.path.getmtime)
print(f'Parsing: {os.path.basename(newest)}')

wb = openpyxl.load_workbook(newest)
ws = wb.active

campaign_name = ws.cell(1, 1).value or ''
period = ws.cell(2, 2).value or ''

records = []
for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
    if not row[0]:
        continue
    price_str = str(row[9] or '0').replace(',', '')
    records.append({
        'date': row[3] or '',
        'time': row[4] or '',
        'batch': row[5] or '',
        'productGroup': row[6] or '',
        'price': int(price_str) if price_str.isdigit() else 0,
        'distributor': row[13] or '',
        'status': row[33] or '',
        'pickupDate': row[34] or '',
        'pickupTime': row[35] or '',
    })

wb.close()

data = {
    'name': campaign_name,
    'period': period,
    'updatedAt': datetime.now().isoformat(timespec='seconds'),
    'records': records
}

out_path = os.path.join(os.path.dirname(__file__) or '.', 'data.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, separators=(',', ':'))

size_kb = os.path.getsize(out_path) // 1024
print(f'Done: {len(records):,} records -> data.json ({size_kb}KB)')

# Git commit & push
try:
    cwd = os.path.dirname(__file__) or '.'
    subprocess.run(['git', 'add', 'data.json'], cwd=cwd, check=True)
    msg = f'Update data: {len(records):,} records ({datetime.now().strftime("%Y-%m-%d %H:%M")})'
    subprocess.run(['git', 'commit', '-m', msg], cwd=cwd, check=True)
    subprocess.run(['git', 'push'], cwd=cwd, check=True)
    print('Deployed! Dashboard will update in ~1 minute.')
except Exception as e:
    print(f'Git push failed: {e}')
    print('Data file is ready. Please git add/commit/push manually.')
